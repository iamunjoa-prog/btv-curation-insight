import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="콘텐츠 블록 주간 실적", layout="wide", page_icon="📊")

st.markdown("""
<style>
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }
.main-title { font-size:24px; font-weight:800; color:#1a1a2e; margin-bottom:2px; }
.sub-title   { font-size:12px; color:#888; margin-bottom:18px; }
.sec-hdr {
    font-size:13px; font-weight:700; color:#1a1a2e;
    border-left:4px solid #4f6ef7; padding-left:9px; margin:24px 0 10px;
}
.sec-hdr.g { border-color:#059669; }

/* pivot table */
.ptable { width:100%; border-collapse:collapse; font-size:12px; }
.ptable th {
    background:#1c1c1e; color:#fff; font-weight:600; font-size:11px;
    padding:9px 10px; text-align:center; white-space:nowrap; border:none;
}
.ptable th.lft { text-align:left; }
.ptable th.wow { background:#2d3a6b; }
.ptable td { padding:7px 10px; border-bottom:1px solid #f2f2f7; }
.ptable td.r   { text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
.ptable td.lbl { font-weight:700; color:#1c1c1e; background:#e8e8ed; }
.ptable td.lbl-num { font-weight:700; color:#1c1c1e; background:#f2f2f7;
                     text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
.ptable td.lbl-wow { font-weight:700; color:#1c1c1e; background:#e8eaf6;
                     text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
.ptable td.sub { color:#3c3c43; padding-left:22px; }
.ptable td.wow-cell { background:#f0f3ff; text-align:right;
                      font-variant-numeric:tabular-nums; white-space:nowrap; }
.ptable tr:hover td { background:#f9f9fb; }
.ptable tr:hover td.lbl { background:#dde0e8; }
.pos { color:#1d4ed8; font-weight:700; }
.neg { color:#dc2626; font-weight:700; }

/* rank table */
.rtable { width:100%; border-collapse:collapse; font-size:12px; }
.rtable th {
    background:#1c1c1e; color:#fff; font-size:11px; font-weight:600;
    padding:8px 10px; text-align:center; white-space:nowrap;
}
.rtable th.lft { text-align:left; }
.rtable td { padding:6px 10px; border-bottom:1px solid #f2f2f7; }
.rtable td.r   { text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums; }
.rtable td.ctr { text-align:center; }
.rtable tr:nth-child(even) td { background:#fafafa; }
.rtable tr:hover td { background:#f0f4ff; }
.new-tag { color:#9333ea; font-weight:700; }
</style>
""", unsafe_allow_html=True)


# ── helpers ──────────────────────────────────────────────────────────

def get_week_label(date):
    """
    엑셀 수식과 동일한 주차 계산:
    월요일+3일(목요일)이 속한 달/주차 기준
    예) 3/30(월) → 목요일 4/2 → 2026-04-W1
        4/20(월) → 목요일 4/23 → 2026-04-W4
        5/4(월)  → 목요일 5/7  → 2026-05-W1
    """
    monday   = date - pd.Timedelta(days=date.weekday())
    thursday = monday + pd.Timedelta(days=3)

    year  = thursday.year
    month = thursday.month

    # 그 달의 첫 번째 목요일 찾기
    first_day = pd.Timestamp(year=year, month=month, day=1)
    first_thu = first_day + pd.Timedelta(days=(3 - first_day.weekday()) % 7)

    week_num = ((thursday - first_thu).days // 7) + 1
    return f"{year}-{month:02d}-W{week_num}"

def get_week_period(dates):
    lo, hi = dates.min(), dates.max()
    return f"{lo.month}/{lo.day}~{hi.month}/{hi.day}"

def fmt_n(v):
    try: return f"{int(round(float(v))):,}"
    except: return "-"

def fmt_wow(cur, prev):
    """WoW 절대값 증감 (+ 파랑 / △ 빨강)"""
    if prev is None or pd.isna(prev) or prev == 0:
        return ""
    d = cur - prev
    if d == 0: return ""
    val = f"{int(round(abs(d))):,}"
    if d > 0: return f'<span class="pos">+{val}</span>'
    return f'<span class="neg">△{val}</span>'

def fmt_wow_pct(cur, prev):
    """WoW % (+ 파랑 / △ 빨강)"""
    if prev is None or pd.isna(prev) or prev == 0:
        return ""
    p = (cur - prev) / prev * 100
    if abs(p) < 0.05: return ""
    val = f"{abs(p):.0f}%"
    if p > 0: return f'<span class="pos">+{val}</span>'
    return f'<span class="neg">△{val}</span>'

def fmt_rank_wow(cur, prev, is_new=False):
    """랭킹용 전주대비 %
    - is_new=True : 전주에 블록 자체 없음 → 모든 지표 NEW
    - prev=0, cur>0 : 증감% (0→X는 NEW 아님, 그냥 계산)
    - prev=0, cur=0 : -
    """
    if is_new:
        return '<span class="new-tag">NEW</span>'
    if prev is None or pd.isna(prev):
        return '<span class="new-tag">NEW</span>'
    if prev == 0 and cur == 0:
        return '-'
    if prev == 0:
        return '-'  # 전주 0이면 % 계산 불가 → -
    p = (cur - prev) / prev * 100
    if abs(p) < 0.05: return '-'
    val = f"{abs(p):.0f}%"
    if p > 0: return f'<span class="pos">+{val}</span>'
    return f'<span class="neg">△{val}</span>'

def load_data(race_file, gen_file):
    """
    Race 파일, 일반 파일 각각 업로드
    - 3행 헤더 (header=2), 4행부터 데이터
    - 열 순서 고정: A=일자, B=블록명, C=UV, D=PV, E=매출 (헤더명 무관)
    """
    COL_NAMES = ["일자", "블록명", "UV", "PV", "매출"]

    def read_one(f):
        # header=None, skiprows=3 → 헤더명 완전 무시, 4행부터 읽기
        raw = pd.read_excel(f, sheet_name=0, header=None, skiprows=3, engine='openpyxl')
        df  = raw.iloc[:, :5].copy()
        df.columns = COL_NAMES
        # 숫자 변환
        df["UV"]   = pd.to_numeric(df["UV"],  errors="coerce").fillna(0)
        df["PV"]   = pd.to_numeric(df["PV"],  errors="coerce").fillna(0)
        df["매출"]  = pd.to_numeric(df["매출"], errors="coerce").fillna(0)
        # 날짜 변환 (빈 행 제거)
        df = df[df["일자"].notna()].copy()
        df["일자"] = pd.to_datetime(df["일자"].astype(str).str[:8], format="%Y%m%d", errors="coerce")
        df = df[df["일자"].notna()].copy()
        # 블록명 정리
        df["블록명"] = df["블록명"].astype(str).str.strip()
        df = df[~df["블록명"].isin(["nan", "", "#"])].copy()
        # 주차 레이블
        df["week_label"] = df["일자"].apply(get_week_label)
        return df.reset_index(drop=True)

    return read_one(race_file), read_one(gen_file)

def calc_weekly(df):
    rows = []
    for week, g in df.groupby('week_label', sort=True):
        daily = g.groupby('일자').agg(UV=('UV','sum'), PV=('PV','sum'), 매출=('매출','sum'))
        rows.append({
            'week':      week,
            'period':    get_week_period(g['일자']),
            'op_count':  g['블록명'].nunique(),
            'avg_uv':    daily['UV'].mean(),
            'avg_pv':    daily['PV'].mean(),
            'total_매출': g['매출'].sum(),
            'total_uv':  g['UV'].sum(),
        })
    return pd.DataFrame(rows)

def calc_rankings(df):
    out = {}
    for week, g in df.groupby('week_label', sort=True):
        def block_stats(x):
            return pd.Series({
                '일평균UV': x.groupby('일자')['UV'].sum().mean(),
                '일평균PV': x.groupby('일자')['PV'].sum().mean(),
                '총매출':   x['매출'].sum(),
                '총UV':    x['UV'].sum(),
            })
        stats = (
            g.groupby('블록명')
             .apply(block_stats, include_groups=False)
             .reset_index()
             .sort_values('일평균UV', ascending=False)
             .reset_index(drop=True)
        )
        out[week] = stats
    return out


# ── Summary pivot HTML ────────────────────────────────────────────────

def render_pivot(race_w, gen_w, all_weeks):
    rm = {r['week']: r for _, r in race_w.iterrows()} if race_w is not None else {}
    gm = {r['week']: r for _, r in gen_w.iterrows()}

    def total_uv(w):
        return rm.get(w, {}).get('total_uv', 0) + gm.get(w, {}).get('total_uv', 0)

    def total_pv_sum(w):
        # 전체 PV: 일평균 PV * 일수 (근사)
        r = rm.get(w, {})
        g = gm.get(w, {})
        return r.get('avg_pv', 0) + g.get('avg_pv', 0)

    # 헤더
    html = '<table class="ptable"><thead><tr>'
    html += '<th class="lft" style="min-width:110px;">구분</th>'
    html += '<th class="lft" style="min-width:90px;">상세 구분</th>'
    for w in all_weeks:
        p = rm.get(w, gm.get(w, {})).get('period','')
        html += f'<th>{w}<br><span style="font-weight:400;font-size:10px;opacity:.7;">({p})</span></th>'
    html += '<th class="wow">WoW</th><th class="wow">WoW(%)</th></tr></thead><tbody>'

    def data_row(label, sub, vals, is_group=False):
        nonlocal html
        if is_group:
            # 전체 행: A·B열 진한 배경, 숫자열 중간 배경, WoW열 별도 배경
            html += f'<tr><td class="lbl">{label}</td><td class="lbl">{sub}</td>'
            for v in vals:
                html += f'<td class="lbl-num">{fmt_n(v) if v is not None else "-"}</td>'
            valid = [v for v in vals if v is not None]
            if len(valid) >= 2:
                html += f'<td class="lbl-wow">{fmt_wow(valid[-1], valid[-2])}</td>'
                html += f'<td class="lbl-wow">{fmt_wow_pct(valid[-1], valid[-2])}</td>'
            else:
                html += '<td class="lbl-wow"></td><td class="lbl-wow"></td>'
        else:
            html += f'<tr><td class="sub">{label}</td><td class="sub">{sub}</td>'
            for v in vals:
                html += f'<td class="r">{fmt_n(v) if v is not None else "-"}</td>'
            valid = [v for v in vals if v is not None]
            if len(valid) >= 2:
                html += f'<td class="wow-cell">{fmt_wow(valid[-1], valid[-2])}</td>'
                html += f'<td class="wow-cell">{fmt_wow_pct(valid[-1], valid[-2])}</td>'
            else:
                html += '<td class="wow-cell"></td><td class="wow-cell"></td>'
        html += '</tr>'

    def get_vals(key, src='race'):
        result = []
        for w in all_weeks:
            d = rm.get(w, {}) if src == 'race' else gm.get(w, {})
            result.append(d.get(key, None))
        return result

    def combined_vals(key):
        result = []
        for w in all_weeks:
            r = rm.get(w, {}).get(key, 0) or 0
            g = gm.get(w, {}).get(key, 0) or 0
            result.append(r + g)
        return result

    # ── 운영 수량
    data_row('운영 수량', '전체', combined_vals('op_count'), is_group=True)
    data_row('', '일반', get_vals('op_count', 'gen'))
    data_row('', 'RACE', get_vals('op_count', 'race'))

    # ── 일평균 UV
    total_avg_uv = []
    for w in all_weeks:
        r = rm.get(w, {}).get('avg_uv', 0) or 0
        g = gm.get(w, {}).get('avg_uv', 0) or 0
        total_avg_uv.append(r + g)
    data_row('일평균 UV', '전체', total_avg_uv, is_group=True)
    data_row('', '일반', get_vals('avg_uv', 'gen'))
    data_row('', 'RACE', get_vals('avg_uv', 'race'))

    # ── 일평균 PV
    total_avg_pv = []
    for w in all_weeks:
        r = rm.get(w, {}).get('avg_pv', 0) or 0
        g = gm.get(w, {}).get('avg_pv', 0) or 0
        total_avg_pv.append(r + g)
    data_row('일평균 PV', '전체', total_avg_pv, is_group=True)
    data_row('', '일반', get_vals('avg_pv', 'gen'))
    data_row('', 'RACE', get_vals('avg_pv', 'race'))

    # ── UV→매출 (원/UV)
    def conv_vals(src):
        result = []
        for w in all_weeks:
            d = rm.get(w, {}) if src == 'race' else gm.get(w, {})
            uv = d.get('total_uv', 0) or 0
            매출 = d.get('total_매출', 0) or 0
            result.append(매출 / uv if uv else None)
        return result

    total_conv = []
    for w in all_weeks:
        r = rm.get(w, {}); g = gm.get(w, {})
        uv = (r.get('total_uv',0) or 0) + (g.get('total_uv',0) or 0)
        매출 = (r.get('total_매출',0) or 0) + (g.get('total_매출',0) or 0)
        total_conv.append(매출 / uv if uv else None)
    data_row('UV→매출', '전체', total_conv, is_group=True)
    data_row('', '일반', conv_vals('gen'))
    data_row('', 'RACE', conv_vals('race'))

    html += '</tbody></table>'
    return html


# ── Ranking HTML ──────────────────────────────────────────────────────

def render_ranking(rankings, week, prev_week, label, color):
    rank_df  = rankings.get(week)
    prev_df  = rankings.get(prev_week) if prev_week else None
    if rank_df is None:
        return "<p>데이터 없음</p>"

    html = f'<table class="rtable"><thead><tr>'
    html += '<th style="width:40px;">Rank</th>'
    html += '<th class="lft">블록명</th>'
    html += '<th>일평균 UV</th><th>일평균 PV</th><th>매출</th>'
    html += '<th>전주대비 UV%</th><th>전주대비 PV%</th><th>전주대비 매출%</th>'
    html += '</tr></thead><tbody>'

    for idx, rr in rank_df.iterrows():
        prev_uv = prev_pv = prev_매출 = None
        # 전주에 블록 자체가 있었는지 UV 기준으로 판단
        is_new = True
        if prev_df is not None:
            m = prev_df[prev_df['블록명'] == rr['블록명']]
            if not m.empty:
                is_new   = False  # 전주에 블록 존재
                prev_uv  = m.iloc[0]['일평균UV']
                prev_pv  = m.iloc[0]['일평균PV']
                prev_매출 = m.iloc[0]['총매출']

        html += '<tr>'
        html += f'<td class="ctr" style="font-weight:700;color:{color};">{idx+1}</td>'
        html += f'<td>{rr["블록명"]}</td>'
        html += f'<td class="r">{fmt_n(rr["일평균UV"])}</td>'
        html += f'<td class="r">{fmt_n(rr["일평균PV"])}</td>'
        html += f'<td class="r">{fmt_n(rr["총매출"])}</td>'
        html += f'<td class="r">{fmt_rank_wow(rr["일평균UV"], prev_uv, is_new)}</td>'
        html += f'<td class="r">{fmt_rank_wow(rr["일평균PV"], prev_pv, is_new)}</td>'
        html += f'<td class="r">{fmt_rank_wow(rr["총매출"],   prev_매출, is_new)}</td>'
        html += '</tr>'

    html += '</tbody></table>'
    return html


# ── Excel export ──────────────────────────────────────────────────────

def export_excel(race_w, gen_w, race_r, gen_r, all_weeks):
    from openpyxl.styles import GradientFill

    wb = Workbook()

    # ── 공통 스타일 ──
    th = Side(style='thin', color='E5E5EA')
    BD = Border(left=th, right=th, top=th, bottom=th)
    NO_BORDER = Border()
    C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    R  = Alignment(horizontal='right',  vertical='center')
    L  = Alignment(horizontal='left',   vertical='center', indent=1)

    dark    = PatternFill('solid', fgColor='1C1C1E')
    gray_bg = PatternFill('solid', fgColor='F2F2F7')
    white   = PatternFill('solid', fgColor='FFFFFF')
    race_bg = PatternFill('solid', fgColor='EEF2FF')
    gen_bg  = PatternFill('solid', fgColor='ECFDF5')
    pos_font = Font(color='1D4ED8', bold=True, size=10)
    neg_font = Font(color='DC2626', bold=True, size=10)
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    lbl_font = Font(bold=True, size=10, color='1C1C1E')
    sub_font = Font(size=10, color='3C3C43')
    num_font = Font(size=10)

    NUM = '#,##0'

    def set_cell(ws, row, col, val, fill=None, font=None, align=None,
                 fmt=None, border=None):
        c = ws.cell(row=row, column=col, value=val)
        if fill:   c.fill      = fill
        if font:   c.font      = font
        if align:  c.alignment = align
        if fmt:    c.number_format = fmt
        c.border = border if border is not None else BD
        return c

    def wow_val(cur, prev):
        if prev is None or prev == 0: return None, None
        d = cur - prev
        p = d / prev * 100
        return d, p

    rm = {r['week']: r for _, r in race_w.iterrows()} if race_w is not None else {}
    gm = {r['week']: r for _, r in gen_w.iterrows()}

    # ════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Summary'
    ws.freeze_panes = 'C3'

    # Title row
    last_col = 2 + len(all_weeks) + 2
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=last_col)
    tc = ws.cell(row=1, column=1,
                 value='콘텐츠 블록 주간 실적 Summary')
    tc.font      = Font(bold=True, size=14, color='1C1C1E')
    tc.fill      = PatternFill('solid', fgColor='F2F2F7')
    tc.alignment = C
    ws.row_dimensions[1].height = 30

    # Header row
    ws.cell(row=2, column=1, value='구분').fill = dark
    ws.cell(row=2, column=1).font = hdr_font
    ws.cell(row=2, column=1).alignment = C
    ws.cell(row=2, column=1).border = BD

    ws.cell(row=2, column=2, value='상세 구분').fill = dark
    ws.cell(row=2, column=2).font = hdr_font
    ws.cell(row=2, column=2).alignment = C
    ws.cell(row=2, column=2).border = BD

    for wi, week in enumerate(all_weeks):
        p = rm.get(week, gm.get(week, {})).get('period','')
        c = ws.cell(row=2, column=3+wi, value=f'{week}\n({p})')
        c.fill = dark; c.font = hdr_font
        c.alignment = C; c.border = BD
        ws.row_dimensions[2].height = 30

    wow_col  = 3 + len(all_weeks)
    wowp_col = wow_col + 1
    for col, lbl in [(wow_col,'WoW'), (wowp_col,'WoW(%)')]:
        c = ws.cell(row=2, column=col, value=lbl)
        c.fill = dark; c.font = hdr_font
        c.alignment = C; c.border = BD

    # Data rows helper
    def write_pivot_row(ws, row_num, label, sub, vals,
                        is_group=False, fill=None):
        bg = fill or (gray_bg if is_group else white)
        fn = lbl_font if is_group else sub_font
        ind = '' if is_group else '  '

        set_cell(ws, row_num, 1, label, bg, fn, L)
        set_cell(ws, row_num, 2, f'{ind}{sub}', bg, fn, L)

        for ci, v in enumerate(vals):
            num = round(float(v)) if v is not None else None
            c = set_cell(ws, row_num, 3+ci, num, bg, num_font, R,
                         fmt=NUM)

        # WoW
        valid = [v for v in vals if v is not None]
        if len(valid) >= 2:
            cur, prv = valid[-1], valid[-2]
            d, p = wow_val(cur, prv)
            if d is not None:
                wc = set_cell(ws, row_num, wow_col, round(d), bg, None, R, NUM)
                wc.font = pos_font if d >= 0 else neg_font
                wp = set_cell(ws, row_num, wowp_col,
                              round(p, 1)/100, bg, None, R, '0%')
                wp.font = pos_font if p >= 0 else neg_font
            else:
                set_cell(ws, row_num, wow_col,  None, bg, num_font, R)
                set_cell(ws, row_num, wowp_col, None, bg, num_font, R)
        else:
            set_cell(ws, row_num, wow_col,  None, bg, num_font, R)
            set_cell(ws, row_num, wowp_col, None, bg, num_font, R)

        ws.row_dimensions[row_num].height = 16

    def gv(key, src, w):
        d = rm.get(w,{}) if src=='race' else gm.get(w,{})
        return d.get(key, None)

    def combined(key):
        return [(rm.get(w,{}).get(key,0) or 0) + (gm.get(w,{}).get(key,0) or 0)
                for w in all_weeks]

    r = 3
    # 운영 수량
    write_pivot_row(ws, r, '운영 수량', '전체', combined('op_count'), True)
    write_pivot_row(ws, r+1, '', '일반', [gv('op_count','gen',w) for w in all_weeks])
    write_pivot_row(ws, r+2, '', 'RACE', [gv('op_count','race',w) for w in all_weeks])
    r += 3

    # 일평균 UV
    tot_uv = [(rm.get(w,{}).get('avg_uv',0) or 0)+(gm.get(w,{}).get('avg_uv',0) or 0)
              for w in all_weeks]
    write_pivot_row(ws, r, '일평균 UV', '전체', tot_uv, True)
    write_pivot_row(ws, r+1, '', '일반', [gv('avg_uv','gen',w) for w in all_weeks])
    write_pivot_row(ws, r+2, '', 'RACE', [gv('avg_uv','race',w) for w in all_weeks])
    r += 3

    # 일평균 PV
    tot_pv = [(rm.get(w,{}).get('avg_pv',0) or 0)+(gm.get(w,{}).get('avg_pv',0) or 0)
              for w in all_weeks]
    write_pivot_row(ws, r, '일평균 PV', '전체', tot_pv, True)
    write_pivot_row(ws, r+1, '', '일반', [gv('avg_pv','gen',w) for w in all_weeks])
    write_pivot_row(ws, r+2, '', 'RACE', [gv('avg_pv','race',w) for w in all_weeks])
    r += 3

    # UV→매출
    def conv(src, w):
        d = rm.get(w,{}) if src=='race' else gm.get(w,{})
        uv = d.get('total_uv',0) or 0
        m  = d.get('total_매출',0) or 0
        return m/uv if uv else None

    tot_conv = []
    for w in all_weeks:
        uv = (rm.get(w,{}).get('total_uv',0) or 0) + (gm.get(w,{}).get('total_uv',0) or 0)
        m  = (rm.get(w,{}).get('total_매출',0) or 0) + (gm.get(w,{}).get('total_매출',0) or 0)
        tot_conv.append(m/uv if uv else None)

    write_pivot_row(ws, r, 'UV→매출(원/UV)', '전체', tot_conv, True)
    write_pivot_row(ws, r+1, '', '일반', [conv('gen',w) for w in all_weeks])
    write_pivot_row(ws, r+2, '', 'RACE', [conv('race',w) for w in all_weeks])

    # Col widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    for wi in range(len(all_weeks)):
        ws.column_dimensions[get_column_letter(3+wi)].width = 15
    ws.column_dimensions[get_column_letter(wow_col)].width  = 12
    ws.column_dimensions[get_column_letter(wowp_col)].width = 10

    # ════════════════════════════════════════════════════
    # Sheet 2~: Weekly ranking sheets
    # ════════════════════════════════════════════════════
    for wi, week in enumerate(all_weeks):
        prev_week = all_weeks[wi-1] if wi > 0 else None
        ws2 = wb.create_sheet(title=week)

        period = race_w[race_w['week']== week]['period'].values if race_w is not None else []
        if not len(period):
            period = gen_w[gen_w['week']== week]['period'].values
        period_str = period[0] if len(period) else ''

        # Title
        ws2.merge_cells('A1:N1')
        tc2 = ws2.cell(row=1, column=1,
                       value=f'{week}  ({period_str})  블록 실적 랭킹')
        tc2.font = Font(bold=True, size=13, color='1C1C1E')
        tc2.fill = PatternFill('solid', fgColor='F2F2F7')
        tc2.alignment = C
        ws2.row_dimensions[1].height = 28

        rank_hdrs = ['Rank','블록명','일평균 UV','일평균 PV','매출',
                     '전주대비 UV%','전주대비 PV%','전주대비 매출%']

        for section, rankings, bg, color_hex in [
            ('일반 콘텐츠 블록', gen_r,  gen_bg,  '059669'),
            ('RACE 블록',       race_r, race_bg, '3B5BDB'),
        ]:
            if rankings is None:
                continue
            # Section header
            r2 = ws2.max_row + 1
            ws2.merge_cells(start_row=r2, start_column=1,
                            end_row=r2, end_column=len(rank_hdrs))
            sh = ws2.cell(row=r2, column=1, value=section)
            sh.fill = PatternFill('solid', fgColor=color_hex)
            sh.font = Font(bold=True, color='FFFFFF', size=11)
            sh.alignment = L
            ws2.row_dimensions[r2].height = 22

            # Headers
            r2 += 1
            for ci, h in enumerate(rank_hdrs, 1):
                c = ws2.cell(row=r2, column=ci, value=h)
                c.fill = dark; c.font = hdr_font
                c.alignment = C if ci != 2 else L
                c.border = BD
            ws2.row_dimensions[r2].height = 16

            rank_df  = rankings.get(week)
            prev_df  = rankings.get(prev_week) if prev_week else None
            if rank_df is None:
                continue

            for idx, rr in rank_df.iterrows():
                r2 += 1
                prev_uv = prev_pv = prev_m = None
                is_new = True
                if prev_df is not None:
                    m = prev_df[prev_df['블록명'] == rr['블록명']]
                    if not m.empty:
                        is_new  = False
                        prev_uv = m.iloc[0]['일평균UV']
                        prev_pv = m.iloc[0]['일평균PV']
                        prev_m  = m.iloc[0]['총매출']

                row_bg = bg if idx % 2 == 0 else white

                ws2.cell(row=r2, column=1, value=idx+1).alignment = C
                ws2.cell(row=r2, column=1).font  = Font(bold=True, color=color_hex, size=10)
                ws2.cell(row=r2, column=1).fill  = row_bg
                ws2.cell(row=r2, column=1).border = BD

                ws2.cell(row=r2, column=2, value=rr['블록명']).alignment = L
                ws2.cell(row=r2, column=2).font  = Font(size=10)
                ws2.cell(row=r2, column=2).fill  = row_bg
                ws2.cell(row=r2, column=2).border = BD

                for ci, (val, prev_val) in enumerate([
                    (rr['일평균UV'], prev_uv),
                    (rr['일평균PV'], prev_pv),
                    (rr['총매출'],   prev_m),
                ], 3):
                    c = ws2.cell(row=r2, column=ci, value=round(val))
                    c.fill = row_bg; c.font = Font(size=10)
                    c.alignment = R; c.border = BD
                    c.number_format = NUM

                for ci, (val, prev_val) in enumerate([
                    (rr['일평균UV'], prev_uv),
                    (rr['일평균PV'], prev_pv),
                    (rr['총매출'],   prev_m),
                ], 6):
                    if is_new:
                        # 전주에 블록 자체 없음 → 모든 지표 NEW
                        txt = 'NEW'; fn = Font(color='9333EA', bold=True, size=10)
                    elif prev_val is None or prev_val == 0:
                        # 전주 블록 있지만 해당 지표 0 → % 계산 불가
                        txt = '-'; fn = Font(size=10, color='999999')
                    else:
                        p = (val - prev_val) / prev_val * 100
                        if abs(p) < 0.05:
                            txt = '-'; fn = Font(size=10)
                        elif p > 0:
                            txt = f'+{abs(p):.0f}%'
                            fn = Font(color='1D4ED8', bold=True, size=10)
                        else:
                            txt = f'△{abs(p):.0f}%'
                            fn = Font(color='DC2626', bold=True, size=10)
                    c = ws2.cell(row=r2, column=ci, value=txt)
                    c.fill = row_bg; c.font = fn
                    c.alignment = R; c.border = BD

                ws2.row_dimensions[r2].height = 15

            ws2.cell(row=ws2.max_row+1, column=1)  # 빈 행

        # Col widths
        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 30
        for ci in range(3, 9):
            ws2.column_dimensions[get_column_letter(ci)].width = 13

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
#  Main UI
# ═══════════════════════════════════════════════════════════════════════

st.markdown('<div class="main-title">📊 콘텐츠 블록 주간 실적</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">RACE(개인화) · 일반 블록 · 주 단위 자동 정리 툴</div>',
            unsafe_allow_html=True)

# ── 보안 안내 ──────────────────────────────────────────────────────
st.markdown("""
<div style="background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;
            padding:12px 16px;margin:8px 0 16px;font-size:13px;color:#7C3D12;">
    ⚠️ <b>파일 업로드 전 보안 해제 필수</b> &nbsp;|&nbsp;
    엑셀 파일의 <b>읽기 전용 / 보안 경고</b>를 해제한 후 업로드해주세요.<br>
    <span style="color:#92400E;">
    🔒 본 분석 도구는 개인 PC에서만 실행되며, 업로드된 데이터는 외부 서버로 전송되지 않습니다.
    </span>
</div>
""", unsafe_allow_html=True)

# ── 파일 업로드 ────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    st.markdown("**① RACE(개인화) 블록 데이터** <span style='color:#6e6e80;font-size:12px'>선택</span>", unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:12px;color:#6e6e80;line-height:1.8;margin-bottom:6px;">
    Oasis에서 뽑은 파일 그대로 보안 해제 후 업로드<br>
    &nbsp;&nbsp;A열: 기준일자 &nbsp;|&nbsp; B열: 블록명 &nbsp;|&nbsp; C열: UV<br>
    &nbsp;&nbsp;D열: PV &nbsp;|&nbsp; E열: 매출
    </div>
    """, unsafe_allow_html=True)
    race_file = st.file_uploader("RACE 엑셀 업로드 (.xlsx)", type=["xlsx","xls"], key="race")
with col2:
    st.markdown("**② 일반 콘텐츠 블록 데이터** <span style='color:#DC2626;font-size:12px'>필수</span>", unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:12px;color:#6e6e80;line-height:1.8;margin-bottom:6px;">
    Oasis에서 뽑은 파일 그대로 보안 해제 후 업로드<br>
    &nbsp;&nbsp;A열: 기준일자 &nbsp;|&nbsp; B열: 블록명 &nbsp;|&nbsp; C열: UV<br>
    &nbsp;&nbsp;D열: PV &nbsp;|&nbsp; E열: 매출
    </div>
    """, unsafe_allow_html=True)
    gen_file = st.file_uploader("일반 블록 엑셀 업로드 (.xlsx)", type=["xlsx","xls"], key="gen")

if gen_file is None:
    st.markdown("""
    <div style="background:#f0f4ff;border:2px dashed #4f6ef7;border-radius:12px;
                padding:36px;text-align:center;margin-top:12px;">
        <div style="font-size:40px">📂</div>
        <div style="font-size:16px;font-weight:700;color:#4f6ef7;margin:10px 0 4px">
            일반 콘텐츠 블록 파일을 업로드해 주세요</div>
        <div style="font-size:13px;color:#888">
            RACE 파일은 선택사항입니다</div>
        <div style="font-size:11px;color:#bbb;margin-top:6px">
            Oasis 다운로드 파일 &nbsp;·&nbsp; 보안 해제 후 업로드 &nbsp;·&nbsp; 3행 헤더 / 4행부터 데이터</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

try:
    if race_file is not None:
        race_df, general_df = load_data(race_file, gen_file)
        has_race = True
    else:
        # RACE 없이 일반만 로드
        _, general_df = load_data(gen_file, gen_file)  # gen_file 두 번 읽어서 general_df만 사용
        race_df = None
        has_race = False
except Exception as e:
    st.error(f"파일 읽기 오류: {e}\n\n💡 Oasis에서 다운로드한 파일을 보안 해제 후 업로드해주세요.\n컬럼 순서: A:기준일자, B:블록명, C:UV, D:PV, E:매출")
    st.stop()

with st.sidebar:
    st.markdown("### ⚙️ 설정")
    st.markdown("---")
    st.markdown("**📋 데이터 현황**")
    if has_race:
        st.markdown(f"- RACE 블록: **{race_df['블록명'].nunique()}개**")
    else:
        st.markdown("- RACE 블록: **미업로드**")
    st.markdown(f"- 일반 블록: **{general_df['블록명'].nunique()}개**")
    ref_df = race_df if has_race else general_df
    st.markdown(f"- 기간: **{ref_df['일자'].min().strftime('%Y.%m.%d')}"
                f" ~ {ref_df['일자'].max().strftime('%Y.%m.%d')}**")

gen_weekly    = calc_weekly(general_df)
gen_rankings  = calc_rankings(general_df)

if has_race:
    race_weekly   = calc_weekly(race_df)
    race_rankings = calc_rankings(race_df)
    all_weeks = sorted(set(race_df['week_label']) | set(general_df['week_label']))
else:
    race_weekly   = None
    race_rankings = None
    all_weeks = sorted(set(general_df['week_label']))

# Export button
export_buf = export_excel(race_weekly, gen_weekly,
                          race_rankings, gen_rankings, all_weeks)
st.download_button(
    "📥 엑셀로 내보내기", data=export_buf,
    file_name=f"콘텐츠블록_주간실적_{datetime.now().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary")

st.markdown("---")

# ── ① Summary pivot ──────────────────────────────────────────────────
st.markdown('<div class="sec-hdr">① 주간 실적 Summary</div>', unsafe_allow_html=True)
st.markdown(render_pivot(race_weekly, gen_weekly, all_weeks),
            unsafe_allow_html=True)

st.markdown("---")

# ── ② 주차별 랭킹 ────────────────────────────────────────────────────
st.markdown('<div class="sec-hdr">② 주차별 블록 랭킹</div>', unsafe_allow_html=True)

selected_week = st.selectbox("주차 선택", all_weeks, index=len(all_weeks)-1)
prev_week = all_weeks[all_weeks.index(selected_week)-1] \
            if all_weeks.index(selected_week) > 0 else None

# 기간 정보
period_info = gen_weekly[gen_weekly['week']==selected_week]['period'].values
period_str  = period_info[0] if len(period_info) else ''
st.markdown(f"**{selected_week}** &nbsp;<span style='color:#888;font-size:12px;'>({period_str})</span>",
            unsafe_allow_html=True)

if has_race:
    tab_gen, tab_race = st.tabs(["🟢 일반 콘텐츠 블록", "🔵 RACE 블록"])
    with tab_gen:
        st.markdown(
            render_ranking(gen_rankings, selected_week, prev_week,
                           "일반 콘텐츠 블록", "#059669"),
            unsafe_allow_html=True)
    with tab_race:
        st.markdown(
            render_ranking(race_rankings, selected_week, prev_week,
                           "RACE 블록", "#3B5BDB"),
            unsafe_allow_html=True)
else:
    st.markdown(
        render_ranking(gen_rankings, selected_week, prev_week,
                       "일반 콘텐츠 블록", "#059669"),
        unsafe_allow_html=True)

# ── 블록별 트렌드 분석 ─────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="sec-hdr">③ 블록별 주차 트렌드</div>', unsafe_allow_html=True)

import plotly.graph_objects as go

# 블록 선택
all_blocks = sorted(general_df["블록명"].unique().tolist())
if has_race:
    race_blocks = sorted(race_df["블록명"].unique().tolist())
    block_type  = st.radio("블록 타입", ["일반 콘텐츠 블록", "RACE 블록"], horizontal=True)
    block_list  = all_blocks if block_type == "일반 콘텐츠 블록" else race_blocks
    src_df      = general_df if block_type == "일반 콘텐츠 블록" else race_df
else:
    block_list  = all_blocks
    src_df      = general_df

selected_block = st.selectbox("블록 선택", block_list)

# 선택 블록 주차별 집계
def get_block_trend(df, block_name, weeks):
    rows = []
    for w in weeks:
        wdf = df[(df["블록명"] == block_name) & (df["week_label"] == w)]
        if wdf.empty:
            continue
        days = wdf["일자"].nunique()
        rows.append({
            "주차": w,
            "일평균UV":  round(wdf["UV"].sum()   / days, 1),
            "일평균PV":  round(wdf["PV"].sum()   / days, 1),
            "총매출":    int(wdf["매출"].sum()),
        })
    return pd.DataFrame(rows)

trend_df = get_block_trend(src_df, selected_block, all_weeks)

if trend_df.empty:
    st.info("해당 블록의 데이터가 없습니다.")
else:
    # 차트
    metric = st.radio("지표 선택", ["일평균UV", "일평균PV", "총매출"], horizontal=True)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=trend_df["주차"],
        y=trend_df[metric],
        mode="lines+markers+text",
        text=[f"{v:,.0f}" for v in trend_df[metric]],
        textposition="top center",
        marker=dict(size=8, color="#3B5BDB"),
        line=dict(color="#3B5BDB", width=2),
        hovertemplate="%{x}<br>" + metric + ": %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text=f"📈 {selected_block} — {metric} 주차별 추이",
                   font=dict(size=14)),
        xaxis_title="주차",
        yaxis_title=metric,
        yaxis=dict(tickformat=","),
        height=380,
        margin=dict(t=50, b=40, l=40, r=20),
        plot_bgcolor="#FAFAFA",
        paper_bgcolor="#FFFFFF",
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)

    # 수치 테이블
    with st.expander("📋 수치 데이터 보기"):
        display_df = trend_df.copy()
        display_df["일평균UV"] = display_df["일평균UV"].apply(lambda v: f"{v:,.1f}")
        display_df["일평균PV"] = display_df["일평균PV"].apply(lambda v: f"{v:,.1f}")
        display_df["총매출"]   = display_df["총매출"].apply(lambda v: f"{v:,}")
        st.dataframe(display_df, use_container_width=True, hide_index=True)

    # 엑셀 저장
    trend_buf = io.BytesIO()
    with pd.ExcelWriter(trend_buf, engine="openpyxl") as writer:
        trend_df.to_excel(writer, index=False, sheet_name="트렌드")
    trend_buf.seek(0)
    st.download_button(
        label="⬇️ 트렌드 데이터 엑셀 저장",
        data=trend_buf,
        file_name=f"{selected_block}_트렌드.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.markdown("---")
with st.expander("📋 원본 데이터 보기"):
    if has_race:
        t1, t2 = st.tabs(["RACE", "일반블록"])
        with t1:
            st.dataframe(race_df.drop(columns=['week_label']), use_container_width=True)
        with t2:
            st.dataframe(general_df.drop(columns=['week_label']), use_container_width=True)
    else:
        st.dataframe(general_df.drop(columns=['week_label']), use_container_width=True)
